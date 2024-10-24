# Dashboard/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Organization, FoodOrder, Invoice, Expense
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def index(request):
    # Fetch data for charts
    organizations_count = Organization.objects.count()
    food_orders_count = FoodOrder.objects.count()
    expenses_total = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    invoices_total = Invoice.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    return render(request, 'index.html', {
        'organizations_count': organizations_count,
        'food_orders_count': food_orders_count,
        'expenses_total': expenses_total,
        'invoices_total': invoices_total,
    })

# View for organizations page
def organization(request):
    if request.method == 'POST':
        name = request.POST['name']
        contact_person = request.POST['contact_person']
        email = request.POST['email']
        Organization.objects.create(name=name, contact_person=contact_person, email=email)
        messages.success(request, "Organization added successfully!")
        return redirect('organization')

    organizations = Organization.objects.all()
    return render(request, 'organization.html', {'organizations': organizations})

# View for the food orders page
def food_order_view(request):
    if request.method == 'POST':
        organization_id = request.POST['organization']
        people_served = request.POST['people_served']
        food_items = request.POST['food_items']
        total_cost = request.POST['total_cost']
        
        print(f"Organization ID: {organization_id}, People Served: {people_served}, Food Items: {food_items}, Total Cost: {total_cost}")

        # Create the food order
        food_order = FoodOrder.objects.create(
            organization_id=organization_id,
            people_served=people_served,
            food_items=food_items,
            total_cost=total_cost
        )
        
        print(f"Food Order Created: {food_order}")

        messages.success(request, "Food order added successfully!")
        return redirect('invoicing')  # Redirect to the invoicing page after submission

    organizations = Organization.objects.all()
    return render(request, 'food_order.html', {'organizations': organizations})
# View for invoicing page
def invoicing_view(request):
    # Fetch all invoices
    invoices = Invoice.objects.all()
    
    return render(request, 'invoicing.html', {
        'invoices': invoices,
    })

def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if request.method == 'POST':
        invoice.status = request.POST.get('status', invoice.status)
        invoice.save()
        messages.success(request, "Invoice updated successfully!")
        return redirect('invoicing')
    return render(request, 'edit_invoice.html', {'invoice': invoice})

def delete_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    invoice.delete()
    messages.success(request, "Invoice deleted successfully!")
    return redirect('invoicing')

# View for expense page
def expense_view(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        date = request.POST.get('date')
        # Ensure to set the user when creating the expense
        Expense.objects.create(
            user=request.user,  # Add this line
            category=category,
            amount=amount,
            date=date
        )
        return redirect('expense')  # Redirect after successful creation
    # Handle GET request and render the template
    return render(request, 'expense.html', {'expenses': Expense.objects.all()})

# View for reports page
def reports(request):
    return render(request, 'reports.html')

# View for profile settings page
def profile(request):
    return render(request, 'profile.html')

def export_to_excel(request):
    # Create a workbook and add a worksheet
    workbook = Workbook()
    
    # Function to add a sheet with data
    def add_sheet_with_data(sheet_name, headers, data):
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Define header style
        header_font = Font(bold=True, color="FFFFFF")  # White text
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row_num, row_data in enumerate(data, 2):  # Start from the second row
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

    # Add Organizations data
    org_headers = ['Name', 'Contact Person', 'Email']
    org_data = [(org.name, org.contact_person, org.email) for org in Organization.objects.all()]
    add_sheet_with_data('Organizations', org_headers, org_data)

    # Add Food Orders data
    food_order_headers = ['ID', 'Organization', 'People Served', 'Food Items', 'Total Cost']
    food_order_data = [
        (order.id, order.organization.name, order.people_served, order.food_items, order.total_cost)
        for order in FoodOrder.objects.select_related('organization').all()
    ]
    add_sheet_with_data('Food Orders', food_order_headers, food_order_data)

    # Add Invoices data
    invoice_headers = ['Invoice ID', 'Total Amount', 'Status', 'Date']
    invoice_data = [
        (invoice.id, invoice.total_amount, invoice.status, invoice.created_at.replace(tzinfo=None))  # Convert to naive datetime
        for invoice in Invoice.objects.all()
    ]
    add_sheet_with_data('Invoices', invoice_headers, invoice_data)

    # Add Expenses data
    expense_headers = ['Expense ID', 'Category', 'Amount', 'Date']
    expense_data = [
        (expense.id, expense.category, expense.amount, expense.date.strftime('%Y-%m-%d'))  # Convert date to string
        for expense in Expense.objects.all()
    ]
    add_sheet_with_data('Expenses', expense_headers, expense_data)

    # Create a BytesIO object to save the workbook in memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Move to the beginning of the BytesIO object

    # Create an HTTP response with the Excel file
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reports.xlsx"'
    return response
